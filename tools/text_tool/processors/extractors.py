"""
Text Extractors - Extracción de texto de archivos y URLs.

Funciones:
- extract_text_from_file: extrae de .txt, .pdf, .docx, .xlsx, .csv
- extract_text_from_url: extrae de páginas web
"""

import os
import csv
import re
import logging
from pathlib import Path
from typing import Dict, Any

logger = logging.getLogger(__name__)

# Check availability
try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    from docx import Document as DocxDocument
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    from openpyxl import load_workbook
    XLXS_AVAILABLE = True
except ImportError:
    XLXS_AVAILABLE = False

try:
    import requests
    from bs4 import BeautifulSoup
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False


def extract_text_from_file(file_path: str) -> Dict[str, Any]:
    """Extrae texto de archivos .txt, .pdf, .docx, .xlsx, .csv."""
    if not os.path.exists(file_path):
        return {'success': False, 'error': 'Archivo no encontrado'}

    ext = Path(file_path).suffix.lower()

    try:
        if ext == '.txt':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                text = f.read()
            return {'success': True, 'text': text, 'source': file_path}

        elif ext == '.pdf':
            if not PDFPLUMBER_AVAILABLE:
                return {'success': False, 'error': 'pdfplumber no instalado'}

            text = ""
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text() or ""
                    text += page_text + "\n"
            return {'success': True, 'text': text, 'source': file_path}

        elif ext in ['.docx', '.doc']:
            if not DOCX_AVAILABLE:
                return {'success': False, 'error': 'python-docx no instalado'}

            doc = DocxDocument(file_path)
            text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
            return {'success': True, 'text': text, 'source': file_path}

        elif ext in ['.xlsx', '.xls']:
            if not XLXS_AVAILABLE:
                return {'success': False, 'error': 'openpyxl no instalado'}

            wb = load_workbook(file_path, data_only=True)
            text = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) if cell else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
            return {'success': True, 'text': text, 'source': file_path}

        elif ext == '.csv':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                text = ""
                for row in reader:
                    text += " ".join([str(cell) for cell in row if cell]) + "\n"
            return {'success': True, 'text': text, 'source': file_path}

        else:
            return {'success': False, 'error': f'Formato no soportado: {ext}'}

    except Exception as e:
        return {'success': False, 'error': str(e)}


def _validate_url_format(url: str) -> str:
    """Validate and return URL format with proper scheme."""
    url = url.strip()
    if not url.startswith(('http://', 'https://')):
        logger.warning(f"URL sin protocolo válido: {url}")
        return 'https://' + url
    return url


def _build_request_headers(url: str) -> Dict[str, str]:
    """Build HTTP request headers for web scraping."""
    return {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'es-ES,es;q=0.9,en;q=0.8',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
    }


def _parse_html_to_text(response_text: str, url: str) -> str:
    """Parse HTML response and extract clean text."""
    soup = BeautifulSoup(response_text, 'html.parser')

    for script in soup(['script', 'style']):
        script.decompose()

    text = soup.get_text(separator=' ')
    text = re.sub(r'\s+', ' ', text).strip()

    logger.info(f"Texto extraído: {len(text)} caracteres")
    return text


def extract_text_from_url(url: str) -> Dict[str, Any]:
    """Extrae texto de una URL."""
    logger.info(f"Intentando scrapear URL: {url}")

    if not REQUESTS_AVAILABLE:
        logger.error("requests no está instalado")
        return {'success': False, 'error': 'requests no instalado'}

    try:
        url = _validate_url_format(url)
        headers = _build_request_headers(url)

        logger.info(f"Haciendo request a: {url}")
        response = requests.get(url, headers=headers, timeout=15, allow_redirects=True)
        response.raise_for_status()

        logger.info(f"Response received, status: {response.status_code}")
        response.encoding = 'utf-8'

        text = _parse_html_to_text(response.text, url)

        return {'success': True, 'text': text, 'source': url}

    except requests.exceptions.HTTPError as e:
        logger.error(f"HTTP Error: {e}")
        return {'success': False, 'error': f'HTTP {e.response.status_code}'}
    except requests.exceptions.ConnectionError:
        logger.error("Error de conexión - verifica la URL")
        return {'success': False, 'error': 'Error de conexión - verifica la URL'}
    except requests.exceptions.Timeout:
        logger.error("Timeout - la página tardó mucho en responder")
        return {'success': False, 'error': 'Timeout - la página tardó mucho'}
    except Exception as e:
        logger.error(f"Error al scrapear: {e}")
        return {'success': False, 'error': str(e)}