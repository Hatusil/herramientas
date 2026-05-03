"""
TextAnalyzerProcessor: Procesamiento de análisis de texto.

Arquitectura modular para análisis de texto con soporte para múltiples formatos
y plugin architecture para analizadores.

Modules:
    - extractors: Extracción de texto de archivos y URLs
    - cleaners: Limpieza y normalización de texto
    - analyzers: Analizadores de texto (frecuencia, wordcloud, n-grams, etc.)

Para agregar un nuevo analizador:
    1. Crear función en processor.py o nuevo módulo
    2. seguir patrón: analyzeNombre(text, **kwargs) -> Dict[str, Any]
    3. Retornar {'success': bool, 'data': Any, 'error': str}
"""

import os
import re
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional, Callable
from io import BytesIO
from collections import Counter

logger = logging.getLogger(__name__)

# ============================================================
# CONFIGURACIÓN DE LÍMITES Y HELPERS
# ============================================================

# Límites de tamaño de texto
TEXT_SIZE_WARNING = 100_000  # 100KB - mostrar warning
TEXT_SIZE_LIMIT = 500_000    # 500KB - límite máximo
TEXT_SIZE_CHUNK = 50_000     # 50KB - tamaño de chunk para procesamiento

def check_text_size(text: str) -> Dict[str, Any]:
    """
    Verifica el tamaño del texto y retorna información de estado.
    
    Returns:
        dict con keys:
        - is_too_large: bool - si excede el límite máximo
        - needs_warning: bool - si necesita warning (entre warning y límite)
        - size: int - tamaño en caracteres
        - size_mb: float - tamaño en MB
        - word_count: int - número de palabras
        - estimated_time: str - tiempo estimado de procesamiento
    """
    if not text:
        return {
            'is_too_large': False,
            'needs_warning': False,
            'size': 0,
            'size_mb': 0,
            'word_count': 0,
            'estimated_time': '< 1 seg'
        }
    
    size = len(text)
    word_count = len(text.split())
    size_mb = size / (1024 * 1024)
    
    # Estimar tiempo de procesamiento basado en tamaño
    if size < TEXT_SIZE_WARNING:
        time_est = '< 1 seg'
    elif size < TEXT_SIZE_WARNING * 2:
        time_est = '1-5 seg'
    elif size < TEXT_SIZE_LIMIT:
        time_est = '5-15 seg'
    else:
        time_est = '15-30 seg+'
    
    return {
        'is_too_large': size > TEXT_SIZE_LIMIT,
        'needs_warning': TEXT_SIZE_WARNING < size <= TEXT_SIZE_LIMIT,
        'size': size,
        'size_mb': round(size_mb, 2),
        'word_count': word_count,
        'estimated_time': time_est
    }

def process_in_chunks(text: str, analyzer_func: Callable, chunk_size: int = TEXT_SIZE_CHUNK, **kwargs) -> Dict[str, Any]:
    """
    Procesa texto grande en chunks y combina resultados.
    
    Args:
        text: Texto a procesar
        analyzer_func: Función de análisis a aplicar
        chunk_size: Tamaño de cada chunk
        **kwargs: Argumentos para la función de análisis
    
    Returns:
        dict con resultados combinados
    """
    chunks = []
    current_chunk = []
    current_size = 0
    
    # Dividir por párrafos
    paragraphs = re.split(r'\n\s*\n|\n{2,}', text.strip())
    
    for para in paragraphs:
        para_size = len(para)
        if current_size + para_size > chunk_size and current_chunk:
            chunks.append('\n\n'.join(current_chunk))
            current_chunk = []
            current_size = 0
        current_chunk.append(para)
        current_size += para_size
    
    if current_chunk:
        chunks.append('\n\n'.join(current_chunk))
    
    # Procesar cada chunk y combinar
    combined_results = []
    errors = []
    
    for i, chunk in enumerate(chunks):
        try:
            result = analyzer_func(chunk, **kwargs)
            if result.get('success'):
                combined_results.append(result)
            else:
                errors.append(result.get('error', f'Chunk {i+1} falló'))
        except Exception as e:
            errors.append(f'Chunk {i+1}: {str(e)}')
    
    return {
        'chunks_processed': len(chunks),
        'successful': len(combined_results),
        'errors': errors,
        'results': combined_results
    }

try:
    import nltk
    # NLTK para n-grams
    nltk.data.find('tokenizers/punkt')
    NLTK_AVAILABLE = True
except Exception as e:
        logger.warning(f"NLTK not available: {e}")
        NLTK_AVAILABLE = False

try:
    from wordcloud import WordCloud
    WORDCLOUD_AVAILABLE = True
except ImportError:
    WORDCLOUD_AVAILABLE = False

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    from docx import Document as DocxDocument
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    from openpyxl import load_workbook
    XLXS_AVAILABLE = True
except ImportError:
    XLXS_AVAILABLE = False

try:
    import requests
    from bs4 import BeautifulSoup
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

try:
    from sklearn.feature_extraction.text import CountVectorizer
    from sklearn.decomposition import LatentDirichletAllocation
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False


# =============================================================================
# PLUGIN ARCHITECTURE - Registry para analizadores
# =============================================================================
# Patrón para agregar nuevos analizadores:
#   1. Definir función: analyzeNombre(text, **kwargs) -> Dict[str, Any]
#   2. Registrar: ANALYZER_REGISTRY['nombre'] = {'func': Callable, 'requires': List[str], 'returns': str}
#   3. returns puede ser: 'image', 'text', 'data', 'stats'
#
# Ejemplo de analizador futuro:
#   def analyze_sentiment(text: str) -> Dict[str, Any]:
#       '''Análisis de sentimiento.'''
#       # Retorna: {'success': True, 'data': {'positive': 0.6, 'negative': 0.2, 'neutral': 0.2}}
#       pass
#   ANALYZER_REGISTRY['sentiment'] = {'func': analyze_sentiment, 'returns': 'data'}

ANALYZER_REGISTRY: Dict[str, Dict[str, Any]] = {}
"""Registro central de analizadores disponibles."""


def get_analyzer(name: str) -> Optional[Callable]:
    """Obtiene función de análisis por nombre."""
    return ANALYZER_REGISTRY.get(name, {}).get('func')


def list_analyzers() -> List[str]:
    """Lista todos los analizadores registrados."""
    return list(ANALYZER_REGISTRY.keys())


def get_analyzer_info(name: str) -> Optional[Dict[str, Any]]:
    """Obtiene metadata de un analizador."""
    return ANALYZER_REGISTRY.get(name)


# Stop words comunes (español + inglés)
STOP_WORDS = {
    'es': {'el', 'la', 'los', 'las', 'un', 'una', 'unas', 'unos', 'de', 'del', 'al', 'a', 
           'en', 'con', 'por', 'para', 'sin', 'sobre', 'entre', 'y', 'e', 'o', 'u', 'que',
           'como', 'más', 'pero', 'ni', 'si', 'no', 'sí', 'él', 'ella', 'ellos', 'ellas',
           'este', 'esta', 'estos', 'estas', 'ese', 'esa', 'esos', 'esas', 'esto',
           'mi', 'tu', 'su', 'mis', 'tus', 'sus', 'nuestro', 'nuestra', 'nosotros',
           'ser', 'estar', 'hay', 'fue', 'era', 'son', 'son', 'es', 'está', 'han', 'había',
           'lo', 'su', 'al', 'todo', 'toda', 'todos', 'todas', 'poco', 'poca', 'pocos', 'pocas',
           'mucho', 'mucha', 'muchos', 'muchas', 'otro', 'otra', 'otros', 'otras', 'mismo', 'misma'},
    'en': {'the', 'a', 'an', 'and', 'or', 'but', 'if', 'in', 'to', 'of', 'for', 'on', 
           'with', 'at', 'by', 'from', 'as', 'is', 'was', 'are', 'were', 'be', 'been',
           'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'could', 'should',
           'may', 'might', 'must', 'shall', 'can', 'need', 'dare', 'ought', 'used',
           'it', 'he', 'she', 'they', 'we', 'you', 'i', 'me', 'him', 'her', 'us', 'them',
           'this', 'that', 'these', 'those', 'what', 'which', 'who', 'whom', 'whose',
           'my', 'your', 'his', 'her', 'its', 'our', 'their', 'whose',
           'not', 'no', 'yes', 'all', 'any', 'some', 'such', 'no', 'nor', 'only',
           'very', 'just', 'also', 'now', 'then', 'there', 'here', 'when', 'where',
           'each', 'every', 'both', 'few', 'more', 'most', 'other', 'some', 'any',
           'so', 'than', 'too', 'very', 'just', 'even', 'still', 'already', 'yet'}
}


def extract_text_from_file(file_path: str) -> Dict[str, Any]:
    """Extrae texto de archivos .txt, .pdf, .docx, .xlsx, .csv."""
    if not os.path.exists(file_path):
        return {'success': False, 'error': 'Archivo no encontrado'}
    
    ext = Path(file_path).suffix.lower()
    
    try:
        if ext == '.txt':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                text = f.read()
            return {'success': True, 'text': text, 'source': file_path}
        
        elif ext == '.pdf':
            if not PDFPLUMBER_AVAILABLE:
                return {'success': False, 'error': 'pdfplumber no instalado'}
            
            text = ""
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text() or ""
                    text += page_text + "\n"
            return {'success': True, 'text': text, 'source': file_path}
        
        elif ext in ['.docx', '.doc']:
            if not DOCX_AVAILABLE:
                return {'success': False, 'error': 'python-docx no instalado'}
            
            doc = DocxDocument(file_path)
            text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
            return {'success': True, 'text': text, 'source': file_path}
        
        elif ext in ['.xlsx', '.xls']:
            if not XLXS_AVAILABLE:
                return {'success': False, 'error': 'openpyxl no instalado'}
            
            wb = load_workbook(file_path, data_only=True)
            text = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) if cell else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
            return {'success': True, 'text': text, 'source': file_path}
        
        elif ext == '.csv':
            import csv
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                text = ""
                for row in reader:
                    text += " ".join([str(cell) for cell in row if cell]) + "\n"
            return {'success': True, 'text': text, 'source': file_path}
        
        else:
            return {'success': False, 'error': f'Formato no soportado: {ext}'}
    
    except Exception as e:
        return {'success': False, 'error': str(e)}


def extract_text_from_url(url: str) -> Dict[str, Any]:
    """Extrae texto de una URL."""
    logger.info(f"Intentando scrapear URL: {url}")
    
    if not REQUESTS_AVAILABLE:
        logger.error("requests no está instalado")
        return {'success': False, 'error': 'requests no instalado'}
    
    # Validar que la URL tenga protocolo
    url = url.strip()
    if not url.startswith(('http://', 'https://')):
        logger.warning(f"URL sin protocolo válido: {url}")
        url = 'https://' + url
    
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'es-ES,es;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Cache-Control': 'max-age=0',
            'Sec-Ch-Ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Ch-Ua-Platform': '"Windows"',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
        }
        
        logger.info(f"Haciendo request a: {url}")
        response = requests.get(url, headers=headers, timeout=15, allow_redirects=True)
        response.raise_for_status()
        
        logger.info(f"Response received, status: {response.status_code}")
        # Force UTF-8 encoding to avoid garbled characters
        response.encoding = 'utf-8'
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Extraer texto visible
        for script in soup(['script', 'style']):
            script.decompose()
        
        text = soup.get_text(separator=' ')
        text = re.sub(r'\s+', ' ', text).strip()
        
        logger.info(f"Texto extraído: {len(text)} caracteres")
        return {'success': True, 'text': text, 'source': url}
    
    except requests.exceptions.HTTPError as e:
        logger.error(f"HTTP Error: {e}")
        return {'success': False, 'error': f'HTTP {e.response.status_code}'}
    except requests.exceptions.ConnectionError:
        logger.error("Error de conexión - verifica la URL")
        return {'success': False, 'error': 'Error de conexión - verifica la URL'}
    except requests.exceptions.Timeout:
        logger.error("Timeout - la página tardó mucho en responder")
        return {'success': False, 'error': 'Timeout - la página tardó mucho'}
    except Exception as e:
        logger.error(f"Error al scrapear: {e}")
        return {'success': False, 'error': str(e)}


def clean_text(text: str, remove_stopwords: bool = True, languages: List[str] = ['es', 'en'], exclude_words: List[str] = None) -> str:
    """Limpia el texto: minúsculas, remove signos, stopwords."""
    # Minúsculas
    text = text.lower()
    
    # Remove punctuation
    text = re.sub(r'[^\w\s]', ' ', text)
    
    # Remove numbers
    text = re.sub(r'\d+', '', text)
    
    # Normalizar espacios
    text = re.sub(r'\s+', ' ', text).strip()
    
    if remove_stopwords:
        stop = set()
        for lang in languages:
            stop.update(STOP_WORDS.get(lang, set()))
        words = text.split()
        text = ' '.join(w for w in words if w not in stop and len(w) > 2)
    
    # Excluir palabras custom
    if exclude_words:
        exclude_set = set(w.lower() for w in exclude_words)
        words = text.split()
        text = ' '.join(w for w in words if w not in exclude_set)
    
    return text


def _generate_shape_mask(shape: str, width: int, height: int) -> Optional[Any]:
    """
    Generate mask for WordCloud shapes.
    
    Args:
        shape: Shape name - 'rectangle', 'circle', 'heart', 'star'
        width: Width of the mask in pixels
        height: Height of the mask in pixels
        
    Returns:
        PIL Image mask or None for rectangle (default)
    """
    try:
        import numpy as np
        from PIL import Image
        
        shape = shape.lower() if shape else 'rectangle'
        
        if shape == 'rectangle' or shape == 'rectángulo':
            return None  # wordcloud uses no mask for rectangle
            
        elif shape == 'circle' or shape == 'círculo':
            # Create circular mask
            center_x, center_y = width // 2, height // 2
            radius = min(width, height) // 2 - 5
            
            y, x = np.ogrid[:height, :width]
            mask = (x - center_x) ** 2 + (y - center_y) ** 2 <= radius ** 2
            # Invert: 0 in shape area (where words go), 255 for background
            return (1 - mask.astype(np.uint8)) * 255
            
        elif shape == 'heart' or shape == 'corazón':
            # Create heart shape mask
            y, x = np.ogrid[:height, :width]
            center_x = width // 2
            scale_x = width / 100
            scale_y = height / 100
            
            # Heart equation (simplified)
            x_norm = (x - center_x) / scale_x
            y_norm = (y - height * 0.4) / scale_y
            
            # Two circles + triangle approximation
            left_circle = ((x_norm + 30) ** 2 + y_norm ** 2) <= 900
            right_circle = ((x_norm - 30) ** 2 + y_norm ** 2) <= 900
            triangle = (y_norm >= -50) & (y_norm <= 30) & (np.abs(x_norm) <= (30 - y_norm / 2))
            
            mask = (left_circle | right_circle | triangle).astype(np.uint8)
            # Invert: 0 in shape area (where words go), 255 for background
            return (1 - mask) * 255
            
        elif shape == 'star' or shape == 'estrella':
            # Create star shape mask
            center_x, center_y = width // 2, height // 2
            outer_radius = min(width, height) // 2 - 5
            inner_radius = outer_radius * 0.4
            
            # Start from -pi/2 (top = 12 o'clock) and create 10 points
            angles = np.linspace(-np.pi/2, -np.pi/2 + 2 * np.pi, 10, endpoint=False)
            
            # Create polygon points
            points = []
            for i, angle in enumerate(angles):
                if i % 2 == 0:
                    r = outer_radius
                else:
                    r = inner_radius
                px = center_x + r * np.cos(angle)
                py = center_y + r * np.sin(angle)
                points.append([px, py])
            
            # Create mask from polygon - convert to numpy array
            from PIL import ImageDraw, Image
            mask_img = Image.new('L', (width, height), 0)
            draw = ImageDraw.Draw(mask_img)
            draw.polygon([(p[0], p[1]) for p in points], fill=255)
            
            # Convert to numpy array for wordcloud - invert
            import numpy as np
            mask_array = np.array(mask_img)
            # Invert: 0 where shape is (words go there), 255 for background
            return (255 - mask_array)
            
        else:
            return None  # Unknown shape, fallback to rectangle
            
    except Exception as e:
        logger.warning(f"Could not generate shape mask '{shape}': {e}")
        return None


def analyze_wordcloud(
    text: str,
    n_words: int = 100,
    width: int = 800,
    height: int = 400,
    colormap: str = 'viridis',
    margin: int = 10,
    shape: str = 'rectangle'
) -> Dict[str, Any]:
    """
    Genera una WordCloud con opciones de personalización.
    
    Args:
        text: Texto de entrada
        n_words: Número máximo de palabras a mostrar (default: 100)
        width: Ancho de la imagen en píxeles (default: 800)
        height: Alto de la imagen en píxeles (default: 400)
        colormap: Nombre del colormap de matplotlib (default: 'viridis')
                 Opciones: viridis, plasma, inferno, magma, cividis, 
                 blues, greens, reds, oranges, purples, etc.
        margin: Margen entre palabras en píxeles (default: 10)
        shape: Forma de la máscara - 'rectangle', 'circle', 'heart', 'star' (default: 'rectangle')
        
    Returns:
        Dict con 'success', 'image_data' (bytes), 'message'
    """
    if not WORDCLOUD_AVAILABLE:
        return {'success': False, 'error': 'wordcloud no instalado'}
    
    try:
        # Validate colormap - convert to proper case or fallback
        # Map lowercase to proper colormap names
        colormap_map = {
            'viridis': 'viridis', 'plasma': 'plasma', 'inferno': 'inferno',
            'magma': 'magma', 'cividis': 'cividis', 'blues': 'Blues',
            'greens': 'Greens', 'reds': 'Reds', 'oranges': 'Oranges',
            'purples': 'Purples', 'coolwarm': 'coolwarm', 'rdygn': 'RdYlGn',
            'seismic': 'seismic', 'terrain': 'terrain', 'ocean': 'ocean'
        }
        if colormap.lower() in colormap_map:
            colormap = colormap_map[colormap.lower()]
        else:
            logger.warning(f"Invalid colormap '{colormap}', falling back to 'viridis'")
            colormap = 'viridis'
        
        cleaned = clean_text(text, remove_stopwords=True)
        
        # Generate shape mask if needed
        logger.info(f"Generating WordCloud with shape: {shape}, width: {width}, height: {height}")
        mask = _generate_shape_mask(shape, width, height)
        logger.info(f"Mask generated: {type(mask)}")
        
        # Build WordCloud kwargs
        wc_kwargs = {
            'width': width,
            'height': height,
            'background_color': 'white',
            'max_words': n_words,
            'colormap': colormap,
            'prefer_horizontal': 0.7,
            'margin': margin
        }
        
        # Add mask if not rectangle
        if mask is not None:
            logger.info(f"Adding mask to WordCloud: shape={mask.shape}, dtype={mask.dtype}")
            wc_kwargs['mask'] = mask
            logger.info(f"wc_kwargs mask keys: {list(wc_kwargs.keys())}")
        
        logger.info(f"Creating WordCloud with kwargs: {list(wc_kwargs.keys())}")
        
        try:
            wc = WordCloud(**wc_kwargs)
            logger.info(f"WordCloud created successfully")
        except Exception as e:
            logger.error(f"Error creating WordCloud: {e}")
            return {'success': False, 'error': f'Error creating WordCloud: {e}'}
        
        # Check if WordCloud has mask
        has_mask = hasattr(wc, 'mask_') and wc.mask_ is not None
        logger.info(f"WordCloud has mask: {has_mask}")
        
        try:
            wc.generate(cleaned)
        except Exception as e:
            logger.error(f"Error generating WordCloud: {e}")
            return {'success': False, 'error': f'Error generating WordCloud: {e}'}
        
        # Convertir a imagen
        img_buffer = BytesIO()
        wc.to_image().save(img_buffer, format='PNG')
        img_buffer.seek(0)
        
        return {
            'success': True,
            'image_data': img_buffer.getvalue(),
            'message': f'WordCloud con {n_words} palabras'
        }
    
    except Exception as e:
        return {'success': False, 'error': str(e)}


def analyze_frequency(
    text: str,
    n: int = 20,
    remove_stopwords: bool = True,
    exclude_words: List[str] = None,
    already_cleaned: bool = False
) -> Dict[str, Any]:
    """
    Analiza frecuencia de palabras.
    
    Args:
        text: Texto de entrada
        n: Número de palabras más frecuentes a retornar (default: 20, rango: 20-100+)
        remove_stopwords: Si True, elimina stopwords español/inglés (default: True)
        exclude_words: Lista de palabras adicionales a excluir
        already_cleaned: Si True, asume texto ya está limpio
        
    Returns:
        Dict con 'success', 'frequencies' (dict), 'total_words', 'unique_words'
        
    Note:
        Tested and working with n values from 20 to 100+
    """
    cleaned = text if already_cleaned else clean_text(text, remove_stopwords=remove_stopwords, exclude_words=exclude_words)
    words = cleaned.split()
    
    # Contar
    freq = {}
    for word in words:
        freq[word] = freq.get(word, 0) + 1
    
    # Ordenar
    sorted_freq = sorted(freq.items(), key=lambda x: x[1], reverse=True)[:n]
    
    return {
        'success': True,
        'frequencies': dict(sorted_freq),
        'total_words': len(words),
        'unique_words': len(freq)
    }


def analyze_trends(text: str, n_terms: int = 5, n_sections: int = 10) -> Dict[str, Any]:
    """
    Análisis de tendencias: frecuencia de palabras en diferentes secciones del texto.
    Similar a 'Trends' de Voyant.
    """
    try:
        import matplotlib
        matplotlib.use('Agg')  # Sin GUI
        import matplotlib.pyplot as plt
        from collections import Counter
    except ImportError:
        return {'success': False, 'error': 'matplotlib no instalado'}
    
    cleaned = clean_text(text, remove_stopwords=True)
    words = cleaned.split()
    
    if len(words) < n_sections:
        return {'success': False, 'error': 'Texto muy corto para tendencias'}
    
    # Encontrar palabras más frecuentes
    freq = Counter(words)
    top_words = [w for w, _ in freq.most_common(n_terms)]
    
    if not top_words:
        return {'success': False, 'error': 'No hay palabras suficientes'}
    
    # Dividir texto en secciones y contar frecuencia de cada palabra
    section_size = len(words) // n_sections
    trends_data = {word: [] for word in top_words}
    
    for i in range(n_sections):
        start = i * section_size
        end = start + section_size if i < n_sections - 1 else len(words)
        section_words = words[start:end]
        section_freq = Counter(section_words)
        
        for word in top_words:
            trends_data[word].append(section_freq.get(word, 0))
    
    # Crear gráfico con protección
    try:
        fig, ax = plt.subplots(figsize=(10, 5))
        
        x = range(n_sections)
        for word in top_words:
            ax.plot(x, trends_data[word], marker='o', label=word, linewidth=2)
        
        ax.set_xlabel('Sección del texto')
        ax.set_ylabel('Frecuencia')
        ax.set_title('Tendencias de palabras en el texto')
        ax.legend(loc='upper right', fontsize=8)
        ax.grid(True, alpha=0.3)
        
        # Convertir a imagen
        img_buffer = BytesIO()
        plt.tight_layout()
        plt.savefig(img_buffer, format='PNG', dpi=100)
        img_buffer.seek(0)
        
        plt.close(fig)
        
        return {
            'success': True,
            'image_data': img_buffer.getvalue(),
            'top_words': top_words,
            'trends': trends_data
        }
    except Exception as e:
        logger.error(f"Error generating trends chart: {e}")
        return {'success': False, 'error': f'Error al generar gráfico: {str(e)}'}


def analyze_correlations(text: str, n_terms: int = 15) -> Dict[str, Any]:
    """
    Análisis de correlaciones entre términos.
    Muestra qué palabras aparecen juntas.
    """
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import numpy as np
        from collections import Counter
    except ImportError:
        return {'success': False, 'error': 'matplotlib no instalado'}
    
    cleaned = clean_text(text, remove_stopwords=True)
    words = cleaned.split()
    
    if len(words) < 20:
        return {'success': False, 'error': 'Texto muy corto para correlaciones'}
    
    # Encontrar términos más frecuentes
    freq = Counter(words)
    top_words = [w for w, _ in freq.most_common(n_terms)]
    
    if len(top_words) < 3:
        return {'success': False, 'error': 'No hay suficientes palabras'}
    
    # Crear matriz de co-ocurrencia
    n = len(top_words)
    cooccur = np.zeros((n, n))
    
    word_idx = {w: i for i, w in enumerate(top_words)}
    
    # Window de 5 palabras
    window = 5
    for i, word in enumerate(words):
        if word in word_idx:
            for j in range(max(0, i - window), min(len(words), i + window + 1)):
                other = words[j]
                if other in word_idx and other != word:
                    cooccur[word_idx[word]][word_idx[other]] += 1
    
    # Crear heatmap
    fig, ax = plt.subplots(figsize=(10, 8))
    
    im = ax.imshow(cooccur, cmap='YlOrRd', aspect='auto')
    
    ax.set_xticks(range(n))
    ax.set_yticks(range(n))
    ax.set_xticklabels(top_words, rotation=45, ha='right', fontsize=8)
    ax.set_yticklabels(top_words, fontsize=8)
    
    ax.set_title('Correlaciones entre términos')
    
    # Colorbar
    cbar = plt.colorbar(im, ax=ax)
    cbar.set_label('Co-ocurrencia')
    
    # Convertir a imagen con protección
    try:
        img_buffer = BytesIO()
        plt.tight_layout()
        plt.savefig(img_buffer, format='PNG', dpi=100)
        img_buffer.seek(0)
        plt.close(fig)
        
        return {
            'success': True,
            'image_data': img_buffer.getvalue(),
            'terms': top_words,
            'matrix': cooccur.tolist()
        }
    except Exception as e:
        logger.error(f"Error generating correlations chart: {e}")
        return {'success': False, 'error': f'Error al generar gráfico: {str(e)}'}


def analyze_scatter(text: str, n_terms: int = 30) -> Dict[str, Any]:
    """
    Scatter plot de términos basado en frecuencia y posición.
    """
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        from collections import Counter
    except ImportError:
        return {'success': False, 'error': 'matplotlib no instalado'}
    
    cleaned = clean_text(text, remove_stopwords=True)
    words = cleaned.split()
    
    if len(words) < 20:
        return {'success': False, 'error': 'Texto muy corto'}
    
    # Contar frecuencia y posición promedio
    freq = Counter(words)
    position_sum = {}
    position_count = {}
    
    for i, word in enumerate(words):
        if word not in position_sum:
            position_sum[word] = 0
            position_count[word] = 0
        position_sum[word] += i
        position_count[word] += 1
    
    # Calcular métricas
    data = []
    for word in freq:
        if position_count[word] >= 2:
            avg_position = position_sum[word] / position_count[word]
            normalized_pos = avg_position / len(words)  # 0-1
            data.append({
                'word': word,
                'frequency': freq[word],
                'avg_position': normalized_pos
            })
    
    if not data:
        return {'success': False, 'error': 'Datos insuficientes'}
    
    # Ordenar por frecuencia y tomar top
    data = sorted(data, key=lambda x: x['frequency'], reverse=True)[:n_terms]
    
    # Crear scatter plot con protección
    try:
        fig, ax = plt.subplots(figsize=(10, 6))
        
        freqs = [d['frequency'] for d in data]
        positions = [d['avg_position'] for d in data]
        labels = [d['word'] for d in data]
        
        sizes = [f * 20 + 50 for f in freqs]  # Tamaño según frecuencia
        
        scatter = ax.scatter(positions, freqs, s=sizes, alpha=0.6, c=freqs, cmap='viridis')
        
        # Labels
        for i, label in enumerate(labels):
            ax.annotate(label, (positions[i], freqs[i]), fontsize=8, alpha=0.8)
        
        ax.set_xlabel('Posición promedio en el texto (inicio → fin)')
        ax.set_ylabel('Frecuencia')
        ax.set_title('Distribución de términos en el texto')
        ax.grid(True, alpha=0.3)
        
        # Convertir a imagen
        img_buffer = BytesIO()
        plt.tight_layout()
        plt.savefig(img_buffer, format='PNG', dpi=100)
        img_buffer.seek(0)
        plt.close(fig)
        
        return {
            'success': True,
            'image_data': img_buffer.getvalue(),
            'data': data
        }
    except Exception as e:
        logger.error(f"Error generating scatter chart: {e}")
        return {'success': False, 'error': f'Error al generar gráfico: {str(e)}'}


def analyze_stats(text: str) -> Dict[str, Any]:
    """Estadísticas del corpus."""
    words = text.split()
    sentences = re.split(r'[.!?]+', text)
    sentences = [s for s in sentences if s.strip()]
    
    # Calcular
    total_chars = len(text)
    total_words = len(words)
    unique_words = len(set(words))
    avg_word_len = sum(len(w) for w in words) / total_words if total_words > 0 else 0
    avg_sentence_len = total_words / len(sentences) if sentences else 0
    
    return {
        'success': True,
        'total_chars': total_chars,
        'total_words': total_words,
        'unique_words': unique_words,
        'total_sentences': len(sentences),
        'avg_word_length': round(avg_word_len, 2),
        'avg_sentence_length': round(avg_sentence_len, 2),
        'type_token_ratio': round(unique_words / total_words, 4) if total_words > 0 else 0
    }


def analyze_ngrams(text: str, n: int = 2, top_k: int = 20) -> Dict[str, Any]:
    """
    Analiza n-grams (sin dependencia de nltk).
    
    Args:
        text: Texto de entrada
        n: Tamaño del n-gram (2=bigramas, 3=trigramas) (default: 2)
        top_k: Número de n-grams más frecuentes a retornar (default: 20, rango: 20-100+)
        
    Returns:
        Dict con 'success', 'ngrams' (dict), 'n' (tipo de n-gram), 'total' (total encontrado)
        
    Note:
        Tested and working with top_k values from 20 to 100+
    """
    cleaned = clean_text(text, remove_stopwords=True)
    words = cleaned.split()
    
    if len(words) < n:
        return {'success': False, 'error': 'Texto muy corto para n-grams'}
    
    # Generar n-grams manualmente
    ngram_list = []
    for i in range(len(words) - n + 1):
        ngram_list.append(tuple(words[i:i+n]))
    
    # Contar
    freq = Counter(ngram_list)
    
    # Ordenar
    sorted_ngrams = freq.most_common(top_k)
    
    return {
        'success': True,
        'ngrams': {(' '.join(ng)): count for ng, count in sorted_ngrams},
        'n': n,
        'total': len(ngram_list)
    }


def analyze_kwic(text: str, keyword: str, context: int = 5, max_results: int = 20) -> Dict[str, Any]:
    """
    Análisis KWIC (Keyword In Context) - Concordancia de palabras clave en contexto.
    
    Args:
        text: Texto de entrada
        keyword: Palabra clave a buscar
        context: Número de palabras antes/después a mostrar (default: 5, rango: 1-15)
        max_results: Máximo de ocurrencias a mostrar (default: 20)
        
    Returns:
        Dict con 'success', 'data' ([{'before': str, 'keyword': str, 'after': str}]), 'error': str
        
    Note:
        Implementa concordancia estilo Voyant - muestra palabra clave en su contexto.
    """
    if not text or not text.strip():
        return {'success': False, 'error': 'Ingrese texto para analizar', 'data': None}
    
    # Limpiar texto y buscar palabra clave (case-insensitive)
    keyword = keyword.strip().lower()
    if not keyword:
        return {'success': False, 'error': 'Ingrese una palabra clave', 'data': None}
    
    # Procesar texto - mantener estructura de palabras
    words = text.split()
    words_lower = [w.lower() for w in words]
    
    # Encontrar todas las ocurrencias
    concordances = []
    for i, word in enumerate(words_lower):
        if word == keyword:
            # Calcular contexto antes
            start = max(0, i - context)
            before_words = words[start:i]
            before_text = ' '.join(before_words)
            
            # Keyword en texto original
            keyword_orig = words[i]
            
            # Calcular contexto después
            end = min(len(words), i + context + 1)
            after_words = words[i+1:end]
            after_text = ' '.join(after_words)
            
            concordances.append({
                'before': before_text,
                'keyword': keyword_orig,
                'after': after_text
            })
            
            # Limitar resultados
            if len(concordances) >= max_results:
                break
    
    if not concordances:
        return {
            'success': True,
            'data': [],
            'error': 'No se encontraron ocurrencias'
        }
    
    return {
        'success': True,
        'data': concordances,
        'error': ''
    }


def analyze_topics(text: str, n_topics: int = 5, max_iter: int = 10) -> Dict[str, Any]:
    """
    Análisis de Tópicos usando LDA (Latent Dirichlet Allocation).
    
    Args:
        text: Texto de entrada
        n_topics: Número de tópicos a extraer (default: 5, rango: 3-10)
        max_iter: Máximo de iteraciones para LDA (default: 10)
        
    Returns:
        Dict con 'success', 'data' ([{'topic_id': int, 'words': [{'word': str, 'weight': float}]}]), 'error': str
        
    Note:
        Implementa LDA estilo Voyant - extrae tópicos latentes del texto.
    """
    if not SKLEARN_AVAILABLE:
        return {'success': False, 'error': 'scikit-learn no instalado', 'data': None}
    
    if not text or not text.strip():
        return {'success': False, 'error': 'Ingrese texto para analizar', 'data': None}
    
    # Dividir texto en párrafos (separados por líneas en blanco o múltiples saltos)
    paragraphs = re.split(r'\n\s*\n|\n{2,}', text.strip())
    paragraphs = [p.strip() for p in paragraphs if p.strip()]
    
    # Si no hay suficientes párrafos, intentar dividir por oraciones
    if len(paragraphs) < 3:
        # Dividir en oraciones (aproximadamente 5-10 oraciones por "párrafo")
        sentences = re.split(r'[.!?]+', text.strip())
        sentences = [s.strip() for s in sentences if s.strip() and len(s.strip()) > 20]
        
        if len(sentences) >= 6:
            # Crear pseudo-párrafos de 5 oraciones cada uno
            paragraphs = []
            for i in range(0, len(sentences), 5):
                para = ' '.join(sentences[i:i+5])
                if para:
                    paragraphs.append(para)
    
    if len(paragraphs) < 2:
        return {'success': False, 'error': 'Texto insuficiente. Se requieren al menos 2 secciones o párrafos.', 'data': None}
    
    try:
        # Crear vectorizador Bag of Words
        # Usar stopwords en español e inglés
        vectorizer = CountVectorizer(
            max_df=0.95,  # Ignorar términos que aparecen en >95% de docs
            min_df=1,      # Ignorar términos que aparecen en solo 1 doc
            stop_words='english',  # Listas de stopwords de sklearn
            max_features=1000  # Máximo de features
        )
        
        # Crear matriz documento-término
        doc_term_matrix = vectorizer.fit_transform(paragraphs)
        
        # Obtener vocabulario
        feature_names = vectorizer.get_feature_names_out()
        
        # Crear modelo LDA
        lda = LatentDirichletAllocation(
            n_components=n_topics,
            max_iter=max_iter,
            learning_method='online',
            random_state=42,
            n_jobs=-1
        )
        
        # Ajustar modelo
        lda.fit(doc_term_matrix)
        
        # Extraer tópicos
        topics_data = []
        for topic_idx, topic in enumerate(lda.components_):
            # Obtener top palabras del tópico
            top_word_indices = topic.argsort()[:-11:-1]  # Top 10 palabras
            
            words = []
            for word_idx in top_word_indices:
                word = feature_names[word_idx]
                weight = float(topic[word_idx])
                words.append({'word': word, 'weight': weight})
            
            topics_data.append({
                'topic_id': topic_idx,
                'words': words
            })
        
        return {
            'success': True,
            'data': topics_data,
            'error': ''
        }
        
    except Exception as e:
        logger.error(f"LDA error: {e}")
        return {'success': False, 'error': str(e), 'data': None}


def analyze_wordtree(text: str, phrase: str, max_depth: int = 5) -> Dict[str, Any]:
    """
    Análisis WordTree - visualiza relaciones de palabras en estructura de árbol.
    
    Args:
        text: Texto de entrada
        phrase: Frase raíz a buscar
        max_depth: Profundidad máxima del árbol (default: 5, rango: 2-5)
        
    Returns:
        Dict con 'success', 'image_data' (bytes), 'tree' (dict), 'error': str
        
    Note:
        Implementa visualización estilo Voyant - muestra árbol de palabras
        derivadas de la frase raíz.
    """
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import numpy as np
    except ImportError:
        return {'success': False, 'error': 'matplotlib no instalado', 'image_data': None, 'tree': None}
    
    if not text or not text.strip():
        return {'success': False, 'error': 'Ingrese texto para analizar', 'image_data': None, 'tree': None}
    
    phrase = phrase.strip().lower()
    if not phrase:
        return {'success': False, 'error': 'Ingrese una frase para analizar', 'image_data': None, 'tree': None}
    
    # Procesar texto
    words = text.lower().split()
    phrase_words = phrase.split()
    phrase_len = len(phrase_words)
    
    if len(words) < phrase_len + 1:
        return {'success': False, 'error': 'Texto muy corto para analizar', 'image_data': None, 'tree': None}
    
    # Buscar todas las ocurrencias de la frase y collectar siguientes palabras
    # Usar n-grams para encontrar siguiente palabra después de la frase
    from collections import Counter
    
    # Construir árbol usando una función helper en lugar de defaultdict anidado
    def make_node():
        return {'count': 0, 'children': {}}
    
    tree = make_node()
    
    # Encontrar ocurrencias de la frase y colectar palabras siguientes
    phrase_as_tuple = tuple(phrase_words)
    
    for i in range(len(words) - phrase_len):
        # Verificar si las palabras coinciden con la frase
        if tuple(words[i:i+phrase_len]) == phrase_as_tuple:
            # Found - colectar siguientes palabras (hasta max_depth)
            current_level = tree
            current_level['count'] += 1
            
            # Colectar siguientes palabras
            for depth in range(max_depth):
                next_idx = i + phrase_len + depth
                if next_idx >= len(words):
                    break
                
                next_word = words[next_idx]
                
                # Solo añadir si no es empty y no es solo punctiation
                if next_word and len(next_word) > 0:
                    # Add to children
                    if next_word not in current_level['children']:
                        current_level['children'][next_word] = make_node()
                    current_level['children'][next_word]['count'] += 1
                    
                    # Move to next level for next iteration
                    current_level = current_level['children'][next_word]
    
    # Verificar si hay datos en el árbol
    if tree['count'] == 0 and len(tree['children']) == 0:
        return {'success': True, 'data': [], 'error': 'No se encontraron relaciones repetidas', 'tree': {}}
    
    # Convertir árbol a estructura simple para visualización
    def tree_to_dict(node):
        """Convierte árbol a formato simple."""
        result = {
            'count': node.get('count', 0),
            'children': {}
        }
        for child_name, child_data in node.get('children', {}).items():
            result['children'][child_name] = tree_to_dict(child_data)
        return result
    
    tree_data = tree_to_dict(tree)
    
    # Crear visualización de árbol
    try:
        fig, ax = plt.subplots(figsize=(10, 12))
        ax.set_xlim(-0.5, 10.5)
        ax.set_ylim(-0.5, max(8, len(tree['children']) * 1.0 + 1))
        
        # Draw root node (the phrase)
        root_x, root_y = 0.5, 5
        ax.text(root_x, root_y, phrase, fontsize=14, fontweight='bold',
               ha='center', va='center',
               bbox=dict(boxstyle='round,pad=0.3', facecolor='#ADD8E6', edgecolor='black'))
        
        # Draw first level children (next words)
        first_level_items = sorted(tree['children'].items(), key=lambda x: x[1]['count'], reverse=True)
        
        if not first_level_items:
            ax.text(5, 2.5, "No se encontraron relaciones repetidas", 
                   fontsize=12, ha='center', style='italic', color='gray')
        else:
            # Limit first level to 8 items for readability
            first_level_items = first_level_items[:8]
            
            # Calculate spacing
            level_height = 4
            item_spacing = 10 / (len(first_level_items) + 1)
            
            for idx, (word, data) in enumerate(first_level_items):
                x_pos = item_spacing * (idx + 1)
                y_pos = level_height
                
                # Draw line from root
                ax.plot([root_x, x_pos], [root_y - 0.2, y_pos + 0.2], 
                       'k-', linewidth=1, alpha=0.5)
                
                # Draw node
                count_text = f"({data['count']})"
                ax.text(x_pos, y_pos, f"{word}\n{count_text}", 
                       fontsize=10, ha='center', va='center',
                       bbox=dict(boxstyle='round,pad=0.2', facecolor='#E8E8E8', edgecolor='gray'))
            
            # Add continuation indicator if truncated
            if len(tree['children']) > 8:
                ax.text(5, 0.5, "... (continúa)", fontsize=10, 
                       ha='center', style='italic', color='gray')
        
        ax.set_title(f'Árbol de Palabras: "{phrase}"', fontsize=14, fontweight='bold')
        ax.axis('off')
        
        # Convertir a imagen
        img_buffer = BytesIO()
        plt.tight_layout()
        plt.savefig(img_buffer, format='PNG', dpi=100)
        img_buffer.seek(0)
        plt.close(fig)
        
        return {
            'success': True,
            'image_data': img_buffer.getvalue(),
            'tree': tree_data,
            'error': ''
        }
        
    except Exception as e:
        logger.error(f"WordTree visualization error: {e}")
        return {'success': False, 'error': str(e), 'image_data': None, 'tree': None}
def _register_analyzers():
    """Registra todos los analizadores disponibles."""
    global ANALYZER_REGISTRY
    
    ANALYZER_REGISTRY.update({
        'wordcloud': {
            'func': analyze_wordcloud,
            'requires': ['wordcloud'],
            'returns': 'image',
            'description': 'Genera nube de palabras',
            'min_words': 10
        },
        'frequency': {
            'func': analyze_frequency,
            'requires': [],
            'returns': 'text',
            'description': 'Palabras más frecuentes',
            'min_words': 5
        },
        'stats': {
            'func': analyze_stats,
            'requires': [],
            'returns': 'stats',
            'description': 'Estadísticas del corpus',
            'min_words': 1
        },
        'ngrams': {
            'func': analyze_ngrams,
            'requires': [],
            'returns': 'text',
            'description': 'N-grams (bigramas, trigramas)',
            'min_words': 3
        },
        'trends': {
            'func': analyze_trends,
            'requires': ['matplotlib'],
            'returns': 'image',
            'description': 'Tendencia de términos por secciones',
            'min_words': 50
        },
        'correlations': {
            'func': analyze_correlations,
            'requires': ['matplotlib', 'numpy'],
            'returns': 'image',
            'description': 'Co-ocurrencia de términos',
            'min_words': 20
        },
        'scatter': {
            'func': analyze_scatter,
            'requires': ['matplotlib'],
            'returns': 'image',
            'description': 'Distribución término-posición',
            'min_words': 20
        },
        'kwic': {
            'func': analyze_kwic,
            'requires': [],
            'returns': 'data',
            'description': 'KWIC - Keyword In Context',
            'min_words': 50
        },
        'topics': {
            'func': analyze_topics,
            'requires': ['sklearn'],
            'returns': 'data',
            'description': 'LDA - Latent Dirichlet Allocation',
            'min_words': 100
        },
        'wordtree': {
            'func': analyze_wordtree,
            'requires': ['matplotlib'],
            'returns': 'image',
            'description': 'WordTree - Árbol de palabras',
            'min_words': 50
        }
    })


# Inicializar registry al importar
_register_analyzers()


# =============================================================================
# UTILIDADES - Funciones helper para analizadores
# =============================================================================

def check_dependencies(requires: List[str]) -> Dict[str, Any]:
    """
    Verifica si las dependencias están disponibles.
    
    Args:
        requires: Lista de librerías requeridas
        
    Returns:
        {'success': bool, 'missing': List[str], 'error': str}
    """
    missing = []
    
    dep_map = {
        'wordcloud': WORDCLOUD_AVAILABLE,
        'matplotlib': 'matplotlib',
        'numpy': 'numpy',
        'pdfplumber': PDFPLUMBER_AVAILABLE,
        'docx': DOCX_AVAILABLE,
        'requests': REQUESTS_AVAILABLE,
        'sklearn': SKLEARN_AVAILABLE
    }
    
    for req in requires:
        if req in dep_map and not dep_map[req]:
            missing.append(req)
    
    if missing:
        return {'success': False, 'missing': missing, 'error': f'Faltan librerías: {", ".join(missing)}'}
    
    return {'success': True, 'missing': [], 'error': ''}


def get_text_stats(text: str) -> Dict[str, Any]:
    """Obtiene estadísticas básicas del texto sin limpiar."""
    words = text.split()
    sentences = re.split(r'[.!?]+', text)
    sentences = [s for s in sentences if s.strip()]
    
    return {
        'chars': len(text),
        'words': len(words),
        'unique': len(set(words)),
        'sentences': len(sentences),
        'avg_word_len': sum(len(w) for w in words) / len(words) if words else 0
    }


def validate_text(text: str, min_words: int = 1) -> Dict[str, Any]:
    """
    Valida que el texto tenga suficiente contenido.
    
    Args:
        text: Texto a validar
        min_words: Mínimo de palabras requeridas
        
    Returns:
        {'valid': bool, 'error': str, 'word_count': int}
    """
    if not text or not text.strip():
        return {'valid': False, 'error': 'Texto vacío', 'word_count': 0}
    
    words = len(text.split())
    
    if words < min_words:
        return {'valid': False, 'error': f'Texto muy corto (mínimo {min_words} palabras)', 'word_count': words}
    
    return {'valid': True, 'error': '', 'word_count': words}